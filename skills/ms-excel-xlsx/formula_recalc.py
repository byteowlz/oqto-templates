#!/usr/bin/env python3
"""
Excel Workbook Formula Processor
Recalculates all formula expressions in Excel workbooks using LibreOffice
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def initialize_office_macro():
    """Initialize LibreOffice macro for formula processing if not already present"""
    if platform.system() == "Darwin":
        macro_directory = os.path.expanduser(
            "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
        )
    else:
        macro_directory = os.path.expanduser(
            "~/.config/libreoffice/4/user/basic/Standard"
        )

    macro_path = os.path.join(macro_directory, "Module1.xba")

    if os.path.exists(macro_path):
        with open(macro_path, "r") as file_handle:
            if "RecalculateAndSave" in file_handle.read():
                return True

    if not os.path.exists(macro_directory):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
        )
        os.makedirs(macro_directory, exist_ok=True)

    macro_code = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

    try:
        with open(macro_path, "w") as file_handle:
            file_handle.write(macro_code)
        return True
    except Exception:
        return False


def process_workbook(filepath, wait_time=30):
    """
    Process formulas in Excel workbook and identify any errors

    Args:
        filepath: Path to Excel workbook
        wait_time: Maximum duration to wait for processing (seconds)

    Returns:
        dict with error locations and counts
    """
    if not Path(filepath).exists():
        return {"error": f"File {filepath} does not exist"}

    full_path = str(Path(filepath).absolute())

    if not initialize_office_macro():
        return {"error": "Failed to initialize LibreOffice macro"}

    command = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        full_path,
    ]

    # Handle timeout command differences between Linux and macOS
    if platform.system() != "Windows":
        timeout_executable = "timeout" if platform.system() == "Linux" else None
        if platform.system() == "Darwin":
            # Check if gtimeout is available on macOS
            try:
                subprocess.run(
                    ["gtimeout", "--version"],
                    capture_output=True,
                    timeout=1,
                    check=False,
                )
                timeout_executable = "gtimeout"
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        if timeout_executable:
            command = [timeout_executable, str(wait_time)] + command

    process_result = subprocess.run(command, capture_output=True, text=True)

    if (
        process_result.returncode != 0 and process_result.returncode != 124
    ):  # 124 is timeout exit code
        error_text = process_result.stderr or "Unknown error during processing"
        if "Module1" in error_text or "RecalculateAndSave" not in error_text:
            return {"error": "LibreOffice macro not configured properly"}
        else:
            return {"error": error_text}

    # Check for Excel errors in the processed file - scan ALL cells
    try:
        workbook = load_workbook(filepath, data_only=True)

        excel_error_types = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_breakdown = {err: [] for err in excel_error_types}
        error_total = 0

        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            # Check ALL rows and columns - no limits
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_error_types:
                            if err in cell.value:
                                location = f"{sheet}!{cell.coordinate}"
                                error_breakdown[err].append(location)
                                error_total += 1
                                break

        workbook.close()

        # Build result summary
        result_data = {
            "status": "success" if error_total == 0 else "errors_found",
            "total_errors": error_total,
            "error_summary": {},
        }

        # Add non-empty error categories
        for error_type, locations in error_breakdown.items():
            if locations:
                result_data["error_summary"][error_type] = {
                    "count": len(locations),
                    "locations": locations[:20],  # Show up to 20 locations
                }

        # Add formula count for context - also check ALL cells
        workbook_with_formulas = load_workbook(filepath, data_only=False)
        formula_total = 0
        for sheet in workbook_with_formulas.sheetnames:
            worksheet = workbook_with_formulas[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_total += 1
        workbook_with_formulas.close()

        result_data["total_formulas"] = formula_total

        return result_data

    except Exception as ex:
        return {"error": str(ex)}


def run_processor():
    if len(sys.argv) < 2:
        print("Usage: python formula_recalc.py <excel_file> [timeout_seconds]")
        print("\nProcesses all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filepath = sys.argv[1]
    wait_time = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result_data = process_workbook(filepath, wait_time)
    print(json.dumps(result_data, indent=2))


if __name__ == "__main__":
    run_processor()
